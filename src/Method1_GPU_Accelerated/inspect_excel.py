import openpyxl
from pathlib import Path

# GeoMIP/src/Method1_GPU_Accelerated/inspect_excel.py
# parents[0] = Method1_GPU_Accelerated
# parents[1] = src
# parents[2] = GeoMIP
geomip_root = Path(__file__).resolve().parents[2]
excel_path = geomip_root / "results" / "pruebas_Metodo1.xlsx"
print(f"Looking for: {excel_path}")
print(f"Exists: {excel_path.exists()}")

if not excel_path.exists():
    results_dir = geomip_root / "results"
    print(f"\nFiles in {results_dir}:")
    if results_dir.exists():
        for f in results_dir.glob("*.xlsx"):
            print(f"  {f.name}")
    else:
        print(f"  {results_dir} does not exist")
else:
    wb = openpyxl.load_workbook(excel_path)
    print(f"\nSheet names ({len(wb.sheetnames)} total):")
    for i, sheet_name in enumerate(wb.sheetnames):
        ws = wb[sheet_name]
        print(f"  {i}: '{sheet_name}' -> {ws.max_row} rows, {ws.max_column} cols")
        if i < 8:
            first_few = []
            for j in range(1, min(6, ws.max_column + 1)):
                first_few.append(ws.cell(1, j).value)
            print(f"     First row: {first_few}")



