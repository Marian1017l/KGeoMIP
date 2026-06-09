"""
Llena DatosPruebas2026_1.xlsx con resultados de KGeometricSIA y KQNodes.

Ejecutar desde GeoMIP/src/Method2_Dynamic_Programming_Reformulation/:
    python fill_excel.py [--hojas 10A 15B] [--ks 2 3 4 5] [--timeout 300]

KGeometricSIA corre en este proceso (venv GeoMIP).
KQNodes corre en subproceso (venv QNodes).

Salida: docs/DatosPruebas2026_1_resultados.xlsx  (copia con resultados)
"""

import argparse
import json
import subprocess
import sys
import time
import traceback
from pathlib import Path

sys.setrecursionlimit(10000)

import numpy as np
import openpyxl
import pandas as pd

# ── Rutas ─────────────────────────────────────────────────────────────────────

SCRIPT_DIR   = Path(__file__).resolve().parent
GEOMIP_ROOT  = SCRIPT_DIR.parent.parent          # .../GeoMIP
PROJECT_ROOT = GEOMIP_ROOT.parent               # .../projecto-analisis-20261
QNODES_ROOT  = PROJECT_ROOT / "QNodes"
QNODES_PY    = QNODES_ROOT / ".venv" / "bin" / "python"
SAMPLES_GEO  = GEOMIP_ROOT / "data" / "samples"
EXCEL_IN     = PROJECT_ROOT / "docs" / "DatosPruebas2026_1.xlsx"
EXCEL_OUT    = PROJECT_ROOT / "docs" / "DatosPruebas2026_1_resultados.xlsx"

# Columnas en cada hoja (0-indexed):
#   0: #Prueba  1: Alcance  2: Mecanismo
#   k=2: QNodes(3,4,5)  Geo(6,7,8)
#   k=3: QNodes(9,10,11) Geo(12,13,14)
#   k=4: QNodes(15,16,17) Geo(18,19,20)
#   k=5: QNodes(21,22,23) Geo(24,25,26)
K_COL_BASE = {2: 3, 3: 9, 4: 15, 5: 21}  # primera col de cada bloque k

SHEET_CONFIG = {
    "10A": {"sheet": "10A-Elementos",  "n": 10, "pagina": "A", "alphabet": "ABCDEFGHIJ"},
    "15B": {"sheet": "15B-Elementos",  "n": 15, "pagina": "B", "alphabet": "ABCDEFGHIJKLMNO"},
    "20A": {"sheet": "20A-Elementos",  "n": 20, "pagina": "A", "alphabet": "ABCDEFGHIJKLMNOPQRST"},
    "22A": {"sheet": "22A-Elementos",  "n": 22, "pagina": "A", "alphabet": "ABCDEFGHIJKLMNOPQRSTUV"},
    "25A": {"sheet": "25A-Elementos ", "n": 25, "pagina": "A", "alphabet": "ABCDEFGHIJKLMNOPQRSTUVWXY"},
}

# ── Helpers ────────────────────────────────────────────────────────────────────

def to_bits(letras: str, alphabet: str) -> str:
    n = len(alphabet)
    bits = ['0'] * n
    for c in str(letras):
        idx = alphabet.find(c)
        if idx >= 0:
            bits[idx] = '1'
    return ''.join(bits)


def run_kqnodes(estado: str, condicion: str, alcance: str, mecanismo: str,
                k: int, pagina: str, timeout: int) -> dict:
    script = (
        "import sys, os\n"
        f"os.chdir(r'{QNODES_ROOT}')\n"
        f"sys.path.insert(0, r'{QNODES_ROOT}')\n"
        "from src.models.base.application import aplicacion\n"
        f"aplicacion.set_pagina_red_muestra('{pagina}')\n"
        "aplicacion.desactivar_profiling()\n"
        "from src.controllers.manager import Manager\n"
        "from src.strategies.kq_nodes import KQNodes\n"
        "import json, numpy as np\n"
        f"estado    = '{estado}'\n"
        f"condicion = '{condicion}'\n"
        f"alcance   = '{alcance}'\n"
        f"mecanismo = '{mecanismo}'\n"
        f"k = {k}\n"
        "gestor = Manager(estado_inicial=estado)\n"
        "tpm    = gestor.cargar_red()\n"
        "analizador = KQNodes(tpm)\n"
        "sol = analizador.aplicar_estrategia(estado, condicion, alcance, mecanismo, k=k)\n"
        'print(json.dumps({"phi": float(sol.perdida), "tiempo": float(sol.tiempo_ejecucion), "particion": sol.particion}))\n'
    )
    py = str(QNODES_PY) if QNODES_PY.exists() else sys.executable
    try:
        r = subprocess.run([py, "-c", script], capture_output=True, text=True,
                           timeout=timeout, cwd=str(QNODES_ROOT))
        if r.returncode != 0:
            return {"phi": None, "tiempo": None, "particion": None,
                    "error": r.stderr.strip()[-200:]}
        for line in reversed(r.stdout.strip().splitlines()):
            line = line.strip()
            if line.startswith("{"):
                try:
                    return json.loads(line)
                except Exception:
                    pass
        return {"phi": None, "tiempo": None, "particion": None,
                "error": f"no JSON: {r.stdout[-100:]}"}
    except subprocess.TimeoutExpired:
        return {"phi": None, "tiempo": None, "particion": None, "error": f"timeout>{timeout}s"}
    except Exception as exc:
        return {"phi": None, "tiempo": None, "particion": None, "error": str(exc)}


def run_kgeo(estado: str, condicion: str, alcance: str, mecanismo: str,
             k: int, pagina: str, tpm: np.ndarray) -> dict:
    from src.models.base.application import aplicacion
    from src.controllers.manager import Manager
    from src.controllers.strategies.kgeometric import KGeometricSIA
    aplicacion.pagina_sample_network = pagina
    t0 = time.time()
    try:
        sol = KGeometricSIA(Manager(estado_inicial=estado)).aplicar_estrategia(
            condicion, alcance, mecanismo, tpm, k=k)
        return {
            "phi":      float(sol.perdida),
            "tiempo":   time.time() - t0,
            "particion": sol.particion,
            "error":    None,
        }
    except Exception as exc:
        return {"phi": None, "tiempo": None, "particion": None, "error": str(exc)[:100]}


# ── Lógica por hoja ────────────────────────────────────────────────────────────

def procesar_hoja(wb_out, nombre: str, cfg: dict, ks: list, timeout: int):
    sheet_name = cfg["sheet"]
    n = cfg["n"]
    pagina = cfg["pagina"]
    alphabet = cfg["alphabet"]
    estado = "1" + "0" * (n - 1)
    condicion = "1" * n

    tpm_path = SAMPLES_GEO / f"N{n}{pagina}.csv"
    if not tpm_path.exists():
        print(f"  [SKIP] TPM no encontrada: {tpm_path}")
        return

    print(f"\n{'='*70}")
    print(f"  Hoja: {sheet_name}  (N={n}, pagina={pagina})")
    print(f"  TPM: {tpm_path}")
    print(f"  k={ks}  timeout={timeout}s")
    print(f"{'='*70}")

    from src.models.base.application import aplicacion
    aplicacion.profiler_habilitado = False
    aplicacion.pagina_sample_network = pagina
    tpm = np.genfromtxt(str(tpm_path), delimiter=",")

    df = pd.read_excel(str(EXCEL_IN), sheet_name=sheet_name, header=None)
    data = df.iloc[5:55, [0, 1, 2]].copy()
    data.columns = ["prueba", "alcance", "mecanismo"]
    data = data.dropna(subset=["prueba"])

    ws = wb_out[sheet_name] if sheet_name in wb_out.sheetnames else None
    if ws is None:
        print(f"  [WARN] Hoja '{sheet_name}' no existe en el archivo de salida.")
        return

    # Data rows start at Excel row 6 (1-indexed) = ws row 6
    DATA_ROW_OFFSET = 6  # fila Excel donde empieza prueba 1

    print(f"  {'p':>3}  {'alcance':<15}  {'mecanismo':<15}", end="")
    for k in ks:
        print(f"  geo_k{k}     kq_k{k} ", end="")
    print()
    print(f"  {'-'*3}  {'-'*15}  {'-'*15}", end="")
    for k in ks:
        print(f"  {'-'*9}  {'-'*9}", end="")
    print()

    for df_idx, row in data.iterrows():
        p = int(row["prueba"])
        alc_s = str(row["alcance"]).strip()
        mec_s = str(row["mecanismo"]).strip()
        alc = to_bits(alc_s, alphabet)
        mec = to_bits(mec_s, alphabet)

        excel_row = DATA_ROW_OFFSET + (p - 1)  # row number in openpyxl (1-indexed)

        print(f"  {p:>3}  {alc_s:<15}  {mec_s:<15}", end="", flush=True)

        for k in ks:
            base = K_COL_BASE[k]  # 0-indexed col of QNodes Partición

            # KGeometricSIA
            geo = run_kgeo(estado, condicion, alc, mec, k, pagina, tpm)
            geo_phi = f"{geo['phi']:.6f}" if geo["phi"] is not None else f"ERR:{(geo['error'] or '')[:15]}"
            geo_t   = f"{geo['tiempo']:.3f}" if geo["tiempo"] is not None else "N/A"
            geo_par = geo["particion"] or ""

            # KQNodes
            kq = run_kqnodes(estado, condicion, alc, mec, k, pagina, timeout)
            kq_phi = f"{kq['phi']:.6f}" if kq.get("phi") is not None else f"ERR:{(kq.get('error') or '')[:15]}"
            kq_t   = f"{kq['tiempo']:.3f}" if kq.get("tiempo") is not None else "N/A"
            kq_par = kq.get("particion") or ""

            # openpyxl cols are 1-indexed; base is 0-indexed → +1 for QNodes, +4 for Geo
            # QNodes: base+0=Partición, base+1=Pérdida, base+2=Tiempo (0-indexed)
            # Geo:    base+3=Partición, base+4=Pérdida, base+5=Tiempo (0-indexed)
            ws.cell(row=excel_row, column=base + 1).value = kq_par   # QNodes Partición
            ws.cell(row=excel_row, column=base + 2).value = kq_phi   # QNodes Pérdida
            ws.cell(row=excel_row, column=base + 3).value = kq_t     # QNodes Tiempo
            ws.cell(row=excel_row, column=base + 4).value = geo_par  # Geo Partición
            ws.cell(row=excel_row, column=base + 5).value = geo_phi  # Geo Pérdida
            ws.cell(row=excel_row, column=base + 6).value = geo_t    # Geo Tiempo

            print(f"  {geo_phi:>9}  {kq_phi:>9}", end="", flush=True)

        print()

    print(f"\n  Hoja '{sheet_name}' completada.")


# ── Main ───────────────────────────────────────────────────────────────────────

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--hojas",   nargs="+", default=["15B"],
                    help="Redes a procesar: 10A 15A 15B 20A 22A (default: 15B)")
    ap.add_argument("--ks",      nargs="+", type=int, default=[2, 3, 4, 5])
    ap.add_argument("--timeout", type=int, default=300,
                    help="Timeout en segundos por llamada a KQNodes (default: 300)")
    ap.add_argument("--salida",  default=str(EXCEL_OUT))
    args = ap.parse_args()

    ks = sorted(set(args.ks))

    # Copiar Excel original al de salida
    import shutil
    salida = Path(args.salida)
    shutil.copy2(str(EXCEL_IN), str(salida))
    print(f"Excel copiado a: {salida}")

    wb = openpyxl.load_workbook(str(salida))

    for nombre in args.hojas:
        cfg = SHEET_CONFIG.get(nombre)
        if cfg is None:
            print(f"[WARN] Red '{nombre}' no reconocida. Disponibles: {list(SHEET_CONFIG)}")
            continue
        procesar_hoja(wb, nombre, cfg, ks, args.timeout)

    wb.save(str(salida))
    print(f"\nResultados guardados en: {salida}")


if __name__ == "__main__":
    # Importar KGeometricSIA antes del main para validar entorno
    from src.models.base.application import aplicacion
    aplicacion.profiler_habilitado = False
    from src.controllers.manager import Manager
    from src.controllers.strategies.kgeometric import KGeometricSIA

    main()
